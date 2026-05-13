"""Script para generar data.js desde el archivo Excel Metas2025.xlsx"""
import openpyxl
import json
import sys

def generate(excel_path="Metas2025.xlsx", output_path="data.js"):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        pct_2025 = row[6] if row[6] is not None else 0
        pct_cuatrienio = row[8] if row[8] is not None else 0

        if isinstance(pct_2025, (int, float)):
            pct_2025 = round(pct_2025 * 100, 1)
        else:
            pct_2025 = 0

        if isinstance(pct_cuatrienio, (int, float)):
            pct_cuatrienio = round(pct_cuatrienio * 100, 1)
        else:
            pct_cuatrienio = 0

        if pct_2025 > 100:
            pct_2025 = 100
        if pct_cuatrienio > 100:
            pct_cuatrienio = 100

        record = {
            "id": row[0],
            "codigo": str(row[1] or ""),
            "referencia": str(row[2] or ""),
            "programado2025": row[3] if row[3] is not None else 0,
            "lineaBase": str(row[4] if row[4] is not None else "-"),
            "logro2025": row[5] if row[5] is not None else 0,
            "pctCumplimiento2025": pct_2025,
            "metaCuatrienio": row[7] if row[7] is not None else 0,
            "pctCuatrienio": pct_cuatrienio,
            "secretaria": str(row[9] or ""),
            "dimension": str(row[10] or ""),
            "programa": str(row[11] or ""),
            "objetivo": str(row[12] or ""),
            "sector": str(row[13] or ""),
            "producto": str(row[14] or ""),
            "indicador": str(row[16] or ""),
        }
        records.append(record)

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("// Datos generados automáticamente desde Metas2025.xlsx\n")
        f.write("// Fecha de generación: 2026-05-12\n")
        f.write("const METAS_DATA = ")
        json.dump(records, f, ensure_ascii=False, indent=2)
        f.write(";\n")

    print(f"Generados {len(records)} registros en {output_path}")
    return records

if __name__ == "__main__":
    records = generate()
    # Resumen
    secretarias = {}
    for r in records:
        s = r["secretaria"]
        if s not in secretarias:
            secretarias[s] = {"count": 0, "sum_pct": 0}
        secretarias[s]["count"] += 1
        secretarias[s]["sum_pct"] += r["pctCumplimiento2025"]

    print("\nResumen por secretaría:")
    for s in sorted(secretarias.keys()):
        info = secretarias[s]
        avg = info["sum_pct"] / info["count"]
        print(f"  {s}: {info['count']} metas, promedio {avg:.1f}%")
