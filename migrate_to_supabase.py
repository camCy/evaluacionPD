"""Script para migrar datos de Excel a Supabase"""
import openpyxl
from supabase import create_client, Client
import os

URL = "https://aurtxbfckcdqrhtjufpt.supabase.co"
KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImF1cnR4YmZja2NkcXJodGp1ZnB0Iiwicm9sZSI6ImFub24iLCJpYXQiOjE3Nzg2MjAyMTUsImV4cCI6MjA5NDE5NjIxNX0.RTzh2C2OwCM4DVuHmWZhf-Y6wgcBcumUZclWvXT-C9s"

supabase: Client = create_client(URL, KEY)

def safe_float(val):
    if val is None or val == "" or val == "-": return 0.0
    if isinstance(val, (int, float)): return float(val)
    try:
        # Limpiar posibles errores de formato como '2..000' -> '2000' o '2.000'
        clean_val = str(val).replace("..", ".").replace(",", "")
        return float(clean_val)
    except:
        return 0.0

def migrate(excel_path="Metas2025.xlsx"):
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] is None:
            continue

        pct_2025 = row[6] if row[6] is not None else 0
        pct_cuatrienio = row[8] if row[8] is not None else 0

        pct_2025 = safe_float(pct_2025)
        if pct_2025 <= 1: pct_2025 *= 100 # Si viene como decimal 0.85 -> 85
        
        pct_cuatrienio = safe_float(pct_cuatrienio)
        if pct_cuatrienio <= 1: pct_cuatrienio *= 100

        if pct_2025 > 100: pct_2025 = 100
        if pct_cuatrienio > 100: pct_cuatrienio = 100

        record = {
            "id": row[0],
            "codigo": str(row[1] or ""),
            "referencia": str(row[2] or ""),
            "programado2025": safe_float(row[3]),
            "linea_base": str(row[4] if row[4] is not None else "-"),
            "logro2025": safe_float(row[5]),
            "pct_cumplimiento2025": float(pct_2025),
            "meta_cuatrienio": safe_float(row[7]),
            "pct_cuatrienio": float(pct_cuatrienio),
            "secretaria": str(row[9] or ""),
            "dimension": str(row[10] or ""),
            "programa": str(row[11] or ""),
            "objetivo": str(row[12] or ""),
            "sector": str(row[13] or ""),
            "producto": str(row[14] or ""),
            "indicador": str(row[16] or ""),
        }
        records.append(record)

    print(f"Subiendo {len(records)} registros a Supabase...")
    
    # Borrar datos existentes primero para evitar conflictos de duplicados en la migración inicial
    supabase.table("metas").delete().neq("id", -1).execute()
    
    # Subir en lotes de 50
    for i in range(0, len(records), 50):
        batch = records[i:i+50]
        supabase.table("metas").insert(batch).execute()
        print(f"  Enviados registros {i} a {min(i+50, len(records))}")

    print("Migración completada con éxito.")

if __name__ == "__main__":
    migrate()
